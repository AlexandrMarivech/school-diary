# app.py
from flask import Flask, render_template, request, redirect, url_for, session, make_response, flash, send_file
from flask_sqlalchemy import SQLAlchemy
from werkzeug.security import generate_password_hash, check_password_hash
import csv, io, os, datetime
from openpyxl import Workbook
from openpyxl.chart import BarChart, Reference
from openpyxl.utils import get_column_letter
from datetime import datetime

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Flask & DB config â”€â”€â”€â”€â”€â”€â”€â”€â”€
app = Flask(__name__)

BASE_DIR = os.path.abspath(os.path.dirname(__file__))
INSTANCE_DIR = os.path.join(BASE_DIR, "instance")
os.makedirs(INSTANCE_DIR, exist_ok=True)

app.config["SECRET_KEY"] = os.environ.get("SECRET_KEY", "dev-only-CHANGE-ME")
app.config["SQLALCHEMY_DATABASE_URI"] = "sqlite:///" + os.path.join(INSTANCE_DIR, "data.db")
app.config["SQLALCHEMY_TRACK_MODIFICATIONS"] = False

db = SQLAlchemy(app)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Models â”€â”€â”€â”€â”€â”€â”€â”€â”€
class User(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    username = db.Column(db.String(80), unique=True, nullable=False)
    password_hash = db.Column(db.String(200), nullable=False)
    role = db.Column(db.String(20), nullable=False)  # student / teacher / admin
    fullname = db.Column(db.String(120), nullable=True)

class Subject(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    name = db.Column(db.String(80), unique=True, nullable=False)

class Grade(db.Model):
    id = db.Column(db.Integer, primary_key=True)
    student_id = db.Column(db.Integer, db.ForeignKey('user.id'), nullable=False)
    subject_id = db.Column(db.Integer, db.ForeignKey('subject.id'), nullable=False)
    value = db.Column(db.Integer, nullable=False)
    year = db.Column(db.Integer, nullable=False)
    quarter = db.Column(db.Integer, nullable=False)
    week = db.Column(db.Integer, nullable=True)  # 1..10 (Ğ½ĞµĞ¾Ğ±ÑĞ·Ğ°Ñ‚ĞµĞ»ÑŒĞ½Ğ¾Ğµ)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Helpers â”€â”€â”€â”€â”€â”€â”€â”€â”€
def current_year():
    return datetime.date.today().year


# ĞŸĞ¾Ğ·Ğ²Ğ¾Ğ»ÑĞµÑ‚ Ğ²Ñ‹Ğ·Ñ‹Ğ²Ğ°Ñ‚ÑŒ {{ current_year() }} Ğ¿Ñ€ÑĞ¼Ğ¾ Ğ² ÑˆĞ°Ğ±Ğ»Ğ¾Ğ½Ğ°Ñ…
@app.context_processor
def inject_globals():
    return dict(current_year=current_year, datetime=datetime)



def create_demo_data():
    db.drop_all()
    db.create_all()

    # ĞŸÑ€ĞµĞ´Ğ¼ĞµÑ‚Ñ‹
    s1 = Subject(name="Ğ ÑƒÑÑĞºĞ¸Ğ¹")
    s2 = Subject(name="ĞœĞ°Ñ‚ĞµĞ¼Ğ°Ñ‚Ğ¸ĞºĞ°")
    s3 = Subject(name="Ğ¤Ğ¸Ğ·Ğ¸ĞºĞ°")
    db.session.add_all([s1, s2, s3])
    db.session.commit()

    # ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»Ğ¸
    admin = User(username="admin", password_hash=generate_password_hash("admin123"),
                 role="admin", fullname="ĞĞ´Ğ¼Ğ¸Ğ½Ğ¸ÑÑ‚Ñ€Ğ°Ñ‚Ğ¾Ñ€ Ğ¨ĞºĞ¾Ğ»Ñ‹")
    teacher = User(username="teacher", password_hash=generate_password_hash("teach123"),
                   role="teacher", fullname="Ğ˜Ğ²Ğ°Ğ½ Ğ˜Ğ²Ğ°Ğ½Ğ¾Ğ² (Ğ£Ñ‡Ğ¸Ñ‚ĞµĞ»ÑŒ)")

    students = []
    # Ğ¡Ñ†ĞµĞ½Ğ°Ñ€Ğ¸Ğ¸ Ğ´Ğ»Ñ Ñ€Ğ°Ğ·Ğ½Ñ‹Ñ… Ñ‚Ğ¸Ğ¿Ğ¾Ğ² ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ¾Ğ²
    profiles = {
        "ĞÑ‚Ğ»Ğ¸Ñ‡Ğ½Ğ¸Ğº": [5, 5, 5, 5],
        "Ğ¥Ğ¾Ñ€Ğ¾ÑˆĞ¸ÑÑ‚": [4, 4, 5, 4],
        "Ğ¡Ñ€ĞµĞ´Ğ½ÑÑ‡Ğ¾Ğº": [3, 4, 3, 4],
        "Ğ¢Ñ€Ğ¾ĞµÑ‡Ğ½Ğ¸Ğº": [3, 3, 3, 3],
        "Ğ”Ğ²Ğ¾ĞµÑ‡Ğ½Ğ¸Ğº": [2, 2, 3, 2],
    }

    # Ğ¡Ğ¾Ğ·Ğ´Ğ°Ğ´Ğ¸Ğ¼ 30 ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ¾Ğ², Ñ†Ğ¸ĞºĞ»Ğ¾Ğ¼ Ñ€Ğ°Ğ·Ğ´Ğ°Ğ´Ğ¸Ğ¼ Ğ¸Ğ¼ Ğ¿Ñ€Ğ¾Ñ„Ğ¸Ğ»Ğ¸
    for i in range(1, 31):
        if i <= 5:
            prof = "ĞÑ‚Ğ»Ğ¸Ñ‡Ğ½Ğ¸Ğº"
        elif i <= 12:
            prof = "Ğ¥Ğ¾Ñ€Ğ¾ÑˆĞ¸ÑÑ‚"
        elif i <= 22:
            prof = "Ğ¡Ñ€ĞµĞ´Ğ½ÑÑ‡Ğ¾Ğº"
        elif i <= 28:
            prof = "Ğ¢Ñ€Ğ¾ĞµÑ‡Ğ½Ğ¸Ğº"
        else:
            prof = "Ğ”Ğ²Ğ¾ĞµÑ‡Ğ½Ğ¸Ğº"

        u = User(
            username=f"student{i}",
            password_hash=generate_password_hash("stud123"),
            role="student",
            fullname=f"Ğ£Ñ‡ĞµĞ½Ğ¸Ğº {i} ({prof})"
        )
        db.session.add(u)
        students.append((u, profiles[prof]))

    db.session.add_all([admin, teacher])
    db.session.commit()

    # Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğ¾Ñ†ĞµĞ½ĞºĞ¸
    for u, pattern in students:
        for subj in [s1, s2, s3]:
            for q, val in enumerate(pattern, start=1):
                g = Grade(student_id=u.id, subject_id=subj.id, value=val,
                          year=current_year(), quarter=q)
                db.session.add(g)
    db.session.commit()

    print("Demo data created! Users:")
    print("  admin/admin123, teacher/teach123, student1..30/stud123")


# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Auth â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/")
def index():
    if "user_id" in session:
        return redirect(url_for("dashboard"))
    return redirect(url_for("login"))

@app.route("/login", methods=["GET", "POST"])
def login():
    error = ""
    if request.method == "POST":
        username = request.form["username"].strip()
        password = request.form["password"].strip()
        user = User.query.filter_by(username=username).first()
        if user and check_password_hash(user.password_hash, password):
            session["user_id"] = user.id
            session["role"] = user.role
            session["username"] = user.username
            session["fullname"] = user.fullname
            flash("Ğ’Ñ…Ğ¾Ğ´ Ğ²Ñ‹Ğ¿Ğ¾Ğ»Ğ½ĞµĞ½", "success")
            return redirect(url_for("dashboard"))
        error = "ĞĞµĞ¿Ñ€Ğ°Ğ²Ğ¸Ğ»ÑŒĞ½Ñ‹Ğ¹ Ğ»Ğ¾Ğ³Ğ¸Ğ½ Ğ¸Ğ»Ğ¸ Ğ¿Ğ°Ñ€Ğ¾Ğ»ÑŒ"
        flash(error, "danger")
    return render_template("login.html", error=error)

@app.route("/logout")
def logout():
    session.clear()
    flash("Ğ’Ñ‹Ñ…Ğ¾Ğ´ Ğ²Ñ‹Ğ¿Ğ¾Ğ»Ğ½ĞµĞ½", "info")
    return redirect(url_for("login"))

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Dashboard â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/dashboard")
def dashboard():
    if "user_id" not in session:
        flash("Ğ¢Ñ€ĞµĞ±ÑƒĞµÑ‚ÑÑ Ğ²Ñ…Ğ¾Ğ´", "danger")
        return redirect(url_for("login"))

    role = session.get("role")

    news = [
    {"title": "Ğ—Ğ°Ğ¿ÑƒÑ‰ĞµĞ½Ğ° Ğ¾Ğ»Ğ¸Ğ¼Ğ¿Ğ¸Ğ°Ğ´Ğ°", "desc": "ĞœĞ°Ñ‚ĞµĞ¼Ğ°Ñ‚Ğ¸ĞºĞ° Ğ¸ Ñ€ÑƒÑÑĞºĞ¸Ğ¹ ÑĞ·Ñ‹Ğº.",
     "url": "https://www.gov.kz/memleket/entities/edu?lang=ru",
     "image": "https://picsum.photos/400/200?random=1",
     "badge": {"text": "ğŸ”¥ Ğ“Ğ¾Ñ€ÑÑ‡ĞµĞµ", "class": "hot"}},
    {"title": "ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ñ ÑĞ°Ğ¹Ñ‚Ğ°", "desc": "Ğ”Ğ¾Ğ±Ğ°Ğ²Ğ»ĞµĞ½Ñ‹ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ñ‹ Ğ¸ Ğ²Ñ‹Ğ³Ñ€ÑƒĞ·ĞºĞ° Ğ² Excel.",
     "url": "https://github.com/AlexandrMarivech/school-diary",
     "image": "https://picsum.photos/400/200?random=2",
     "badge": {"text": "ğŸ’¡ ĞĞ±Ğ½Ğ¾Ğ²Ğ»ĞµĞ½Ğ¸Ğµ", "class": "tip"}},
    {"title": "ĞĞ¾Ğ²Ñ‹Ğµ Ğ¿Ñ€Ğ¾ĞµĞºÑ‚Ñ‹", "desc": "Ğ ĞµÑĞ¿ÑƒĞ±Ğ»Ğ¸ĞºĞ°Ğ½ÑĞºĞ¸Ğµ Ğ¸Ğ½Ğ¸Ñ†Ğ¸Ğ°Ñ‚Ğ¸Ğ²Ñ‹ Ğ´Ğ»Ñ ÑˆĞºĞ¾Ğ».",
     "url": "https://bilimland.kz/ru/news-articles",
     "image": "https://picsum.photos/400/200?random=3",
     "badge": {"text": "ğŸ†• ĞĞ¾Ğ²Ğ¾Ğµ", "class": "new"}},
]



    return render_template("dashboard.html", role=role, news=news)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Student â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/student")
def student_page():
    if "user_id" not in session or session.get("role") != "student":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑÑ‚ÑƒĞ´ĞµĞ½Ñ‚Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    student_id = session["user_id"]
    year = int(request.args.get("year", current_year()))

    subjects = Subject.query.all()
    subject_map = {s.id: s.name for s in subjects}

    grades = Grade.query.filter_by(student_id=student_id, year=year).all()

    avg = {}
    for g in grades:
        subjname = subject_map.get(g.subject_id, "")
        avg.setdefault(subjname, []).append(g.value)
    avg = {k: round(sum(v)/len(v), 2) if v else 0 for k, v in avg.items()}

    return render_template("student.html", grades=grades, avg=avg, year=year, subject_map=subject_map)

@app.route("/student/report")
def student_report():
    if "user_id" not in session or session.get("role") != "student":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑÑ‚ÑƒĞ´ĞµĞ½Ñ‚Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    student_id = session["user_id"]
    year = int(request.args.get("year", current_year()))

    subjects = Subject.query.all()
    subject_map = {s.id: s.name for s in subjects}

    grades = Grade.query.filter_by(student_id=student_id, year=year).all()

    subj_avgs = {}
    for g in grades:
        subjname = subject_map.get(g.subject_id, "")
        subj_avgs.setdefault(subjname, []).append(g.value)
    subj_avgs = {k: round(sum(v)/len(v), 2) if v else 0 for k, v in subj_avgs.items()}
    overall = round(sum([g.value for g in grades])/len(grades), 2) if grades else 0

    return render_template("student_report.html", year=year,
                           subject_avgs=subj_avgs, overall_avg=overall, subjects=subjects)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Teacher â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/teacher", methods=["GET", "POST"])
def teacher_page():
    if "user_id" not in session or session.get("role") != "teacher":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑƒÑ‡Ğ¸Ñ‚ĞµĞ»ĞµĞ¹", "danger")
        return redirect(url_for("login"))

    subjects = Subject.query.all()
    students = User.query.filter_by(role="student").all()
    message = ""

    if request.method == "POST":
        subject_id = int(request.form["subject"])
        year = int(request.form["year"])
        quarter = int(request.form["quarter"])
        week = int(request.form.get("week", 1))

        for st in students:
            key = f"student_{st.id}"
            raw = request.form.get(key, "").strip()
            if not raw:
                continue
            try:
                value = int(raw)
            except ValueError:
                continue
            if value < 2 or value > 5:
                continue

            existing = Grade.query.filter_by(
                student_id=st.id, subject_id=subject_id,
                year=year, quarter=quarter, week=week
            ).first()
            if existing:
                existing.value = value
            else:
                db.session.add(Grade(
                    student_id=st.id, subject_id=subject_id, value=value,
                    year=year, quarter=quarter, week=week
                ))
        db.session.commit()
        message = "ĞÑ†ĞµĞ½ĞºĞ¸ ÑĞ¾Ñ…Ñ€Ğ°Ğ½ĞµĞ½Ñ‹."

    # âš¡ Ğ¸ÑĞ¿Ñ€Ğ°Ğ²Ğ»ĞµĞ½Ğ¾: Ğ¿ĞµÑ€ĞµĞ´Ğ°Ñ‘Ğ¼ Ñ„ÑƒĞ½ĞºÑ†Ğ¸Ñ, Ğ° Ğ½Ğµ Ñ‡Ğ¸ÑĞ»Ğ¾
    return render_template("teacher.html",
                           subjects=subjects, students=students,
                           message=message, current_year=current_year)


# âš¡ Ğ½Ğ¾Ğ²Ñ‹Ğ¹ Ğ°Ğ»Ğ¸Ğ°Ñ Ğ´Ğ»Ñ ÑÑ‚Ğ°Ñ€Ñ‹Ñ… ÑÑÑ‹Ğ»Ğ¾Ğº
@app.route("/export/class")
def export_class():
    # Ğ¿Ñ€Ğ¾ÑÑ‚Ğ¾ Ğ¿ĞµÑ€ĞµĞ½Ğ°Ğ¿Ñ€Ğ°Ğ²Ğ»ÑĞµĞ¼ Ğº export_teacher_xlsx
    return redirect(url_for("export_teacher_xlsx", **request.args))



@app.route("/teacher/report")
def teacher_report():
    if "user_id" not in session or session.get("role") != "teacher":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑƒÑ‡Ğ¸Ñ‚ĞµĞ»ĞµĞ¹", "danger")
        return redirect(url_for("login"))

    subject_id = int(request.args.get("subject", 0))
    year = int(request.args.get("year", current_year()))
    period = request.args.get("period", "year")  # quarter1..4, halfyear1/2, year

    subjects = Subject.query.all()
    students = User.query.filter_by(role="student").all()

    if period.startswith("quarter"):
        quarters = [int(period[-1])]
    elif period == "halfyear1":
        quarters = [1, 2]
    elif period == "halfyear2":
        quarters = [3, 4]
    else:
        quarters = [1, 2, 3, 4]

    report_data = []
    for st in students:
        q = Grade.query.filter_by(student_id=st.id, year=year)
        if subject_id != 0:
            q = q.filter_by(subject_id=subject_id)
        q = q.filter(Grade.quarter.in_(quarters))
        grades = [g.value for g in q.all()]
        avg = round(sum(grades)/len(grades), 2) if grades else 0
        report_data.append((st.fullname or st.username, grades, avg))

    return render_template("teacher_report.html",
                           subjects=subjects, subject_id=subject_id,
                           year=year, period=period, report_data=report_data)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Excel exports â”€â”€â”€â”€â”€â”€â”€â”€â”€
def autosize_columns(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            try:
                max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
            except Exception:
                pass
        ws.column_dimensions[col_letter].width = max(12, min(40, max_len + 2))

@app.route("/export/teacher_xlsx")
def export_teacher_xlsx():
    # Ğ£Ñ‡Ğ¸Ñ‚ĞµĞ»ÑŒ/ĞĞ´Ğ¼Ğ¸Ğ½: Ğ²Ñ‹Ğ³Ñ€ÑƒĞ·ĞºĞ° Ğ¿Ğ¾ ĞºĞ»Ğ°ÑÑÑƒ (Ñ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ°Ğ¼Ğ¸ Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚/Ğ³Ğ¾Ğ´/Ñ‡ĞµÑ‚Ğ²ĞµÑ€Ñ‚ÑŒ/Ğ½ĞµĞ´ĞµĞ»Ñ)
    if "user_id" not in session or session.get("role") not in ["teacher", "admin"]:
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑƒÑ‡Ğ¸Ñ‚ĞµĞ»ĞµĞ¹/Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    subject_id = int(request.args.get("subject", 0))
    year = int(request.args.get("year", current_year()))
    quarter = int(request.args.get("quarter", 0))
    week = int(request.args.get("week", 0))
    subject = Subject.query.get(subject_id) if subject_id != 0 else None
    students = User.query.filter_by(role="student").all()

    wb = Workbook()
    ws = wb.active
    ws.title = "ĞÑ‚Ñ‡ĞµÑ‚ ĞºĞ»Ğ°ÑÑĞ°"

    ws.append(["Ğ¤Ğ˜Ğ ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ°", "ĞŸÑ€ĞµĞ´Ğ¼ĞµÑ‚", "ĞŸĞµÑ€Ğ¸Ğ¾Ğ´", "ĞĞµĞ´ĞµĞ»Ñ", "ĞÑ†ĞµĞ½ĞºĞ¸", "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ»"])

    for st in students:
        q = Grade.query.filter_by(student_id=st.id, year=year)
        if quarter != 0:
            q = q.filter_by(quarter=quarter)
        if subject_id != 0:
            q = q.filter_by(subject_id=subject_id)
        if week != 0:
            q = q.filter_by(week=week)
        rows = q.all()
        grades = [g.value for g in rows]
        avg = round(sum(grades)/len(grades), 2) if grades else ""
        subjname = subject.name if subject else "Ğ’ÑĞµ"
        period_str = f"{year}, Q{quarter if quarter else '1-4'}"
        week_str = week if week else "Ğ²ÑĞµ"
        ws.append([st.fullname or st.username, subjname, period_str, week_str, ";".join(map(str, grades)), avg])

    autosize_columns(ws)

    # Ğ”Ğ¸Ğ°Ğ³Ñ€Ğ°Ğ¼Ğ¼Ğ° Ğ¿Ğ¾ ÑÑ€ĞµĞ´Ğ½Ğ¸Ğ¼ (ĞºĞ¾Ğ»Ğ¾Ğ½ĞºĞ° F)
    data_start = 2
    data_end = ws.max_row
    if data_end >= data_start:
        chart = BarChart()
        chart.title = "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ» Ğ¿Ğ¾ ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ°Ğ¼"
        values = Reference(ws, min_col=6, min_row=1, max_row=data_end)  # F1..F*
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)    # A2..A*
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.title = "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ»"
        chart.x_axis.title = "Ğ£Ñ‡ĞµĞ½Ğ¸Ğº"
        ws.add_chart(chart, "H2")

    filepath = os.path.join(INSTANCE_DIR, f"teacher_report_{year}_q{quarter}_w{week}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


@app.route("/export/student_xlsx")
def export_student_xlsx():
    # Ğ¡Ñ‚ÑƒĞ´ĞµĞ½Ñ‚: Ğ»Ğ¸Ñ‡Ğ½Ñ‹Ğ¹ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚ Ñ Ğ´Ğ¸Ğ°Ğ³Ñ€Ğ°Ğ¼Ğ¼Ğ¾Ğ¹ Ğ¿Ğ¾ Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚Ğ°Ğ¼
    if "user_id" not in session or session.get("role") != "student":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ ÑÑ‚ÑƒĞ´ĞµĞ½Ñ‚Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    student_id = session["user_id"]
    year = int(request.args.get("year", current_year()))
    subjects = Subject.query.all()
    subject_map = {s.id: s.name for s in subjects}

    q = Grade.query.filter_by(student_id=student_id, year=year).all()
    subj_avgs = {}
    for g in q:
        name = subject_map.get(g.subject_id, "ĞĞµĞ¸Ğ·Ğ².")
        subj_avgs.setdefault(name, []).append(g.value)

    wb = Workbook()
    ws = wb.active
    ws.title = f"ĞÑ‚Ñ‡Ñ‘Ñ‚ {year}"

    ws.append(["ĞŸÑ€ĞµĞ´Ğ¼ĞµÑ‚", "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ»"])
    for s in subjects:
        vals = subj_avgs.get(s.name, [])
        avg = round(sum(vals)/len(vals), 2) if vals else 0
        ws.append([s.name, avg])

    autosize_columns(ws)

    # Ğ”Ğ¸Ğ°Ğ³Ñ€Ğ°Ğ¼Ğ¼Ğ° Ğ¿Ğ¾ Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚Ğ°Ğ¼
    data_end = ws.max_row
    if data_end >= 2:
        chart = BarChart()
        chart.title = "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ» Ğ¿Ğ¾ Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚Ğ°Ğ¼"
        values = Reference(ws, min_col=2, min_row=1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.title = "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ»"
        chart.x_axis.title = "ĞŸÑ€ĞµĞ´Ğ¼ĞµÑ‚"
        ws.add_chart(chart, "E2")

    filepath = os.path.join(INSTANCE_DIR, f"student_report_{year}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Admin â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/admin", methods=["GET", "POST"])
def admin_page():
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    message = ""
    if request.method == "POST":
        username = request.form["username"].strip()
        fullname = request.form.get("fullname", "").strip()
        password = request.form["password"].strip()
        role = request.form["role"]

        if username and password and len(password) > 4 and role in ["student", "teacher", "admin"]:
            if User.query.filter_by(username=username).first():
                message = "ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ Ñ Ñ‚Ğ°ĞºĞ¸Ğ¼ Ğ»Ğ¾Ğ³Ğ¸Ğ½Ğ¾Ğ¼ ÑƒĞ¶Ğµ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚"
                flash(message, "danger")
            else:
                u = User(
                    username=username,
                    password_hash=generate_password_hash(password),
                    role=role,
                    fullname=fullname
                )
                db.session.add(u)
                db.session.commit()
                message = "ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ ÑĞ¾Ğ·Ğ´Ğ°Ğ½"
                flash(message, "success")
        else:
            message = "ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ (Ğ¿Ğ°Ñ€Ğ¾Ğ»ÑŒ >4 ÑĞ¸Ğ¼Ğ²Ğ¾Ğ»Ğ¾Ğ², ĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½Ğ°Ñ Ñ€Ğ¾Ğ»ÑŒ)"
            flash(message, "danger")

    users = User.query.all()
    return render_template("admin.html", users=users, message=message)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Admin: Ñ€ĞµĞ´Ğ°ĞºÑ‚Ğ¸Ñ€Ğ¾Ğ²Ğ°Ğ½Ğ¸Ğµ Ğ¿Ğ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»Ñ â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/edit_user/<int:user_id>", methods=["POST"])
def edit_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    user = User.query.get_or_404(user_id)

    username = request.form.get("username", "").strip()
    fullname = request.form.get("fullname", "").strip()
    role = request.form.get("role", "").strip()
    password = request.form.get("password", "").strip()

    if username:
        user.username = username
    user.fullname = fullname
    if role in ["student", "teacher", "admin"]:
        user.role = role
    if password and len(password) > 4:
        user.password_hash = generate_password_hash(password)

    db.session.commit()
    flash("ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ Ğ¾Ğ±Ğ½Ğ¾Ğ²Ğ»Ñ‘Ğ½", "success")
    return redirect(url_for("admin_page"))


# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Admin: ÑƒĞ´Ğ°Ğ»ĞµĞ½Ğ¸Ğµ Ğ¿Ğ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»Ñ â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/delete_user/<int:user_id>", methods=["POST"])
def delete_user(user_id):
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    user = User.query.get_or_404(user_id)

    # Ğ—Ğ°Ñ‰Ğ¸Ñ‚Ğ° â€” Ğ½ĞµĞ»ÑŒĞ·Ñ ÑƒĞ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¸ÑÑ‚Ñ€Ğ°Ñ‚Ğ¾Ñ€Ğ°
    if user.role == "admin":
        flash("ĞĞµĞ»ÑŒĞ·Ñ ÑƒĞ´Ğ°Ğ»Ğ¸Ñ‚ÑŒ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¸ÑÑ‚Ñ€Ğ°Ñ‚Ğ¾Ñ€Ğ°!", "danger")
        return redirect(url_for("admin_page"))

    db.session.delete(user)
    db.session.commit()
    flash("ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ ÑƒĞ´Ğ°Ğ»Ñ‘Ğ½", "info")
    return redirect(url_for("admin_page"))


    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    message = ""
    if request.method == "POST":
        username = request.form["username"].strip()
        fullname = request.form.get("fullname", "").strip()
        password = request.form["password"].strip()
        role = request.form["role"]
        if username and password and len(password) > 4 and role in ["student", "teacher", "admin"]:
            if User.query.filter_by(username=username).first():
                message = "ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ Ñ Ñ‚Ğ°ĞºĞ¸Ğ¼ Ğ»Ğ¾Ğ³Ğ¸Ğ½Ğ¾Ğ¼ ÑƒĞ¶Ğµ ÑÑƒÑ‰ĞµÑÑ‚Ğ²ÑƒĞµÑ‚"
                flash(message, "danger")
            else:
                u = User(username=username,
                         password_hash=generate_password_hash(password),
                         role=role, fullname=fullname)
                db.session.add(u)
                db.session.commit()
                message = "ĞŸĞ¾Ğ»ÑŒĞ·Ğ¾Ğ²Ğ°Ñ‚ĞµĞ»ÑŒ ÑĞ¾Ğ·Ğ´Ğ°Ğ½"
                flash(message, "success")
        else:
            message = "ĞĞµĞ²ĞµÑ€Ğ½Ñ‹Ğµ Ğ´Ğ°Ğ½Ğ½Ñ‹Ğµ (Ğ¿Ğ°Ñ€Ğ¾Ğ»ÑŒ >4 ÑĞ¸Ğ¼Ğ²Ğ¾Ğ»Ğ¾Ğ², ĞºĞ¾Ñ€Ñ€ĞµĞºÑ‚Ğ½Ğ°Ñ Ñ€Ğ¾Ğ»ÑŒ)"
            flash(message, "danger")

    users = User.query.all()
    return render_template("admin.html", users=users, message=message)


@app.route("/admin/reports")
def admin_reports():
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    year = int(request.args.get("year", current_year()))
    if year < 2000 or year > current_year() + 1:
        year = current_year()
        flash("Ğ“Ğ¾Ğ´ Ğ¸ÑĞ¿Ñ€Ğ°Ğ²Ğ»ĞµĞ½ Ğ½Ğ° Ñ‚ĞµĞºÑƒÑ‰Ğ¸Ğ¹", "info")

    students = User.query.filter_by(role="student").all()
    subjects = Subject.query.all()
    if not subjects:
        flash("ĞĞµÑ‚ Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚Ğ¾Ğ² Ğ² Ğ±Ğ°Ğ·Ğµ. Ğ—Ğ°Ğ¿ÑƒÑÑ‚Ğ¸Ñ‚Ğµ Ğ¸Ğ½Ğ¸Ñ†Ğ¸Ğ°Ğ»Ğ¸Ğ·Ğ°Ñ†Ğ¸Ñ Ğ‘Ğ”.", "danger")
        return redirect(url_for("admin_page"))

    subject_map = {s.id: s.name for s in subjects}
    report_data = []
    for st in students:
        q = Grade.query.filter_by(student_id=st.id, year=year).all()
        subj_avgs = {}
        for g in q:
            subjname = subject_map.get(g.subject_id, "ĞĞµĞ¸Ğ·Ğ²ĞµÑÑ‚Ğ½Ñ‹Ğ¹")
            subj_avgs.setdefault(subjname, []).append(g.value)
        subj_avgs = {k: round(sum(v)/len(v), 2) if v else 0 for k, v in subj_avgs.items()}
        overall_avg = round(sum([g.value for g in q])/len(q), 2) if q else 0
        report_data.append({
            "student": st.fullname or st.username,
            "subj_avgs": subj_avgs,
            "overall": overall_avg
        })

    return render_template("admin_reports.html",
                           year=year,
                           report_data=report_data,
                           total_students=len(students),
                           subjects=subjects)


# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Admin: ÑĞºÑĞ¿Ğ¾Ñ€Ñ‚ Ğ¾Ñ‚Ñ‡Ñ‘Ñ‚Ğ° Ğ² Excel â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/export/admin_xlsx")
def export_admin_xlsx():
    # ĞĞ´Ğ¼Ğ¸Ğ½: ÑĞ²Ğ¾Ğ´Ğ½Ğ°Ñ Ğ¿Ğ¾ ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ°Ğ¼/Ğ¿Ñ€ĞµĞ´Ğ¼ĞµÑ‚Ğ°Ğ¼ (ÑÑ€ĞµĞ´Ğ½Ğ¸Ğµ Ğ·Ğ° Ğ³Ğ¾Ğ´)
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    year = int(request.args.get("year", current_year()))
    students = User.query.filter_by(role="student").all()
    subjects = Subject.query.all()
    subject_map = {s.id: s.name for s in subjects}

    wb = Workbook()
    ws = wb.active
    ws.title = f"Ğ˜Ñ‚Ğ¾Ğ³Ğ¸ {year}"

    # Ğ—Ğ°Ğ³Ğ¾Ğ»Ğ¾Ğ²ĞºĞ¸
    headers = ["Ğ£Ñ‡ĞµĞ½Ğ¸Ğº"] + [s.name for s in subjects] + ["ĞĞ±Ñ‰Ğ¸Ğ¹ ÑÑ€ĞµĞ´Ğ½Ğ¸Ğ¹"]
    ws.append(headers)

    # Ğ”Ğ°Ğ½Ğ½Ñ‹Ğµ
    for st in students:
        q = Grade.query.filter_by(student_id=st.id, year=year).all()
        subj_avgs = {}
        for g in q:
            name = subject_map.get(g.subject_id, "ĞĞµĞ¸Ğ·Ğ².")
            subj_avgs.setdefault(name, []).append(g.value)

        row = [st.fullname or st.username]
        vals_for_mean = []
        for s in subjects:
            vals = subj_avgs.get(s.name, [])
            avg = round(sum(vals)/len(vals), 2) if vals else ""
            row.append(avg)
            if isinstance(avg, (int, float)):
                vals_for_mean.append(avg)

        overall = round(sum(vals_for_mean)/len(vals_for_mean), 2) if vals_for_mean else ""
        row.append(overall)
        ws.append(row)

    # ĞĞ²Ñ‚Ğ¾ÑˆĞ¸Ñ€Ğ¸Ğ½Ğ°
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            max_len = max(max_len, len(str(cell.value)) if cell.value else 0)
        ws.column_dimensions[col_letter].width = max(12, min(40, max_len + 2))

    # Ğ”Ğ¸Ğ°Ğ³Ñ€Ğ°Ğ¼Ğ¼Ğ° Ğ¿Ğ¾ Ğ¾Ğ±Ñ‰ĞµĞ¼Ñƒ ÑÑ€ĞµĞ´Ğ½ĞµĞ¼Ñƒ
    last_col = ws.max_column
    data_end = ws.max_row
    if data_end >= 2:
        chart = BarChart()
        chart.title = "ĞĞ±Ñ‰Ğ¸Ğ¹ ÑÑ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ» (Ğ¿Ğ¾ ÑƒÑ‡ĞµĞ½Ğ¸ĞºĞ°Ğ¼)"
        values = Reference(ws, min_col=last_col, min_row=1, max_row=data_end)
        cats = Reference(ws, min_col=1, min_row=2, max_row=data_end)
        chart.add_data(values, titles_from_data=True)
        chart.set_categories(cats)
        chart.y_axis.title = "Ğ¡Ñ€ĞµĞ´Ğ½Ğ¸Ğ¹ Ğ±Ğ°Ğ»Ğ»"
        chart.x_axis.title = "Ğ£Ñ‡ĞµĞ½Ğ¸Ğº"
        ws.add_chart(chart, f"{get_column_letter(last_col+2)}2")

    filepath = os.path.join(INSTANCE_DIR, f"admin_report_{year}.xlsx")
    wb.save(filepath)
    return send_file(filepath, as_attachment=True)


    
# â”€â”€â”€â”€â”€â”€â”€â”€â”€ Admin Dashboard â”€â”€â”€â”€â”€â”€â”€â”€â”€
@app.route("/admin/dashboard")
def admin_dashboard():
    if "user_id" not in session or session.get("role") != "admin":
        flash("Ğ”Ğ¾ÑÑ‚ÑƒĞ¿ Ñ‚Ğ¾Ğ»ÑŒĞºĞ¾ Ğ´Ğ»Ñ Ğ°Ğ´Ğ¼Ğ¸Ğ½Ğ¾Ğ²", "danger")
        return redirect(url_for("login"))

    return render_template("admin_dashboard.html")


from datetime import datetime

# Ğ ĞµĞ³Ğ¸ÑÑ‚Ñ€Ğ°Ñ†Ğ¸Ñ Ñ„Ğ¸Ğ»ÑŒÑ‚Ñ€Ğ¾Ğ² Ğ¸ Ğ³Ğ»Ğ¾Ğ±Ğ°Ğ»ÑŒĞ½Ñ‹Ñ… Ñ„ÑƒĞ½ĞºÑ†Ğ¸Ğ¹ Ğ´Ğ»Ñ Jinja2
@app.template_filter('datetime_format')
def datetime_format(value, format="%d.%m.%Y %H:%M"):
    return value.strftime(format)

@app.context_processor
def utility_processor():
    def current_year():
        return datetime.now().year
    return dict(current_year=current_year)

# â”€â”€â”€â”€â”€â”€â”€â”€â”€ CLI â”€â”€â”€â”€â”€â”€â”€â”€â”€
if __name__ == "__main__":
    import sys
    with app.app_context():
        db.create_all()
    if "initdb" in sys.argv:
        with app.app_context():
            create_demo_data()
    else:
        app.run(host="0.0.0.0", port=5000, debug=True)
